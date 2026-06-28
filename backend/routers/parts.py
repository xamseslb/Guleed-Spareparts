"""
Router for reservedeler (Parts) – inkluderer CRUD, søk og bildeopplasting.
"""

import os
import uuid
import shutil
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, status
from sqlalchemy.orm import Session
from PIL import Image as PilImage
from backend.database import get_db
from backend.models.part import Part
from backend.models.order import Order
from backend.models.loan import Loan
from backend.schemas.part import PartCreate, PartUpdate, PartOut
from backend.services.auth_service import get_current_user
from backend.models.user import User, UserRole

router = APIRouter(prefix="/api/parts", tags=["Parts"])

UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "..", "uploads")
MAX_IMAGES = 10
ALLOWED_TYPES = {"image/jpeg", "image/png", "image/webp"}
MAX_FILE_SIZE_MB = 5


# ─── Hent alle varer (med søk og filter) ──────────────────────────────
@router.get("/", response_model=List[PartOut])
def get_parts(
    q: Optional[str] = Query(None, description="Search by name or part number"),
    category: Optional[str] = Query(None),
    car_make: Optional[str] = Query(None),
    car_model: Optional[str] = Query(None),
    low_stock_only: bool = Query(False),
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = db.query(Part)

    if q:
        search_term = f"%{q.lower()}%"
        query = query.filter(
            Part.name.ilike(search_term) | Part.part_number.ilike(search_term)
        )
    if category:
        query = query.filter(Part.category.ilike(f"%{category}%"))
    if low_stock_only:
        query = query.filter(Part.stock_quantity <= Part.low_stock_threshold)

    parts = query.offset(skip).limit(limit).all()

    # Biltype-filter (JSON-felt kan ikke filtreres i SQL på tvers av backends)
    if car_make or car_model:
        filtered = []
        for p in parts:
            for car in (p.compatible_cars or []):
                make_match = not car_make or car_make.lower() in car.get("make", "").lower()
                model_match = not car_model or car_model.lower() in car.get("model", "").lower()
                if make_match and model_match:
                    filtered.append(p)
                    break
        parts = filtered

    return parts


# ─── Hent én vare ─────────────────────────────────────────────────────
# Header order + which columns are required, used for the downloadable template.
# NOTE: this static route must be declared BEFORE "/{part_id}", otherwise
# FastAPI matches "import-template" as a part id and returns a 422.
TEMPLATE_COLUMNS = [
    ("part_number", True, "BR-8690"),
    ("name", True, "Front Brake Pads"),
    ("category", True, "Brakes"),
    ("unit_price", True, 25),
    ("description", False, "Fits Toyota Corolla 2008–2014"),
    ("stock_quantity", False, 42),
    ("low_stock_threshold", False, 5),
    ("location", False, "Shelf B, Row 04"),
    ("ordered_quantity", False, 10),
]


@router.get("/import-template")
def import_template(current_user: User = Depends(get_current_user)):
    """A ready-to-fill Excel workbook. Real .xlsx so columns stay separate
    regardless of the user's regional comma/semicolon settings."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.comments import Comment
    from fastapi.responses import Response

    wb = Workbook()
    ws = wb.active
    ws.title = "parts_template"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    example_rows = [
        ["BR-8690", "Front Brake Pads", "Brakes", 25, "Fits Toyota Corolla 2008–2014", 42, 5, "Shelf B, Row 04", 10],
        ["OIL-5W40-2L", "Engine Oil 5W-40 — 2L", "Oils & Fluids", 18, "Fully synthetic, 2 litre", 30, 6, "Shelf A, Row 01", 0],
    ]

    for col_idx, (field, required, _) in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.comment = Comment("Required" if required else "Optional — may be left blank", "Guleed")
        ws.column_dimensions[cell.column_letter].width = max(14, len(field) + 4)

    for r, example in enumerate(example_rows, start=2):
        for c, value in enumerate(example, start=1):
            ws.cell(row=r, column=c, value=value)

    ws.freeze_panes = "A2"  # keep the header visible while scrolling

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="parts_template.xlsx"'},
    )


@router.get("/{part_id}", response_model=PartOut)
def get_part(part_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    part = db.query(Part).filter(Part.id == part_id).first()
    if not part:
        raise HTTPException(status_code=404, detail="Part not found")
    return part


# ─── Legg til vare ────────────────────────────────────────────────────
@router.post("/", response_model=PartOut, status_code=status.HTTP_201_CREATED)
def create_part(data: PartCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    existing = db.query(Part).filter(Part.part_number == data.part_number).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Part number '{data.part_number}' already exists")

    d = data.model_dump()
    # Fill in blanks so a part can be added with just a part number now and
    # named/categorised/priced later.
    d["name"] = (d.get("name") or "").strip() or d["part_number"]
    d["category"] = (d.get("category") or "").strip() or "Uncategorized"
    d["unit_price"] = d.get("unit_price") or 0.0
    part = Part(**d)
    part.compatible_cars = [c.model_dump() for c in data.compatible_cars]
    part.images = []
    db.add(part)
    db.commit()
    db.refresh(part)
    return part


# ─── Oppdater vare ────────────────────────────────────────────────────
@router.put("/{part_id}", response_model=PartOut)
def update_part(
    part_id: int,
    data: PartUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    part = db.query(Part).filter(Part.id == part_id).first()
    if not part:
        raise HTTPException(status_code=404, detail="Part not found")

    update_data = data.model_dump(exclude_unset=True)
    if "compatible_cars" in update_data and update_data["compatible_cars"] is not None:
        update_data["compatible_cars"] = [c.model_dump() if hasattr(c, "model_dump") else c for c in update_data["compatible_cars"]]

    for field, value in update_data.items():
        setattr(part, field, value)

    db.commit()
    db.refresh(part)
    return part


# ─── Slett vare ───────────────────────────────────────────────────────
@router.delete("/{part_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_part(part_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Only admins can delete parts")
    part = db.query(Part).filter(Part.id == part_id).first()
    if not part:
        raise HTTPException(status_code=404, detail="Part not found")

    # Don't orphan order/loan history – block deletion if the part is referenced
    order_count = db.query(Order).filter(Order.part_id == part_id).count()
    loan_count = db.query(Loan).filter(Loan.part_id == part_id).count()
    if order_count or loan_count:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete part: {order_count} order(s) and {loan_count} loan(s) reference it.",
        )

    # Slett alle tilhørende bilder
    for img_path in (part.images or []):
        full_path = os.path.join(UPLOAD_DIR, os.path.basename(img_path))
        if os.path.exists(full_path):
            os.remove(full_path)
    db.delete(part)
    db.commit()


# ─── Last opp bilde ───────────────────────────────────────────────────
@router.post("/{part_id}/images", response_model=PartOut)
async def upload_image(
    part_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    part = db.query(Part).filter(Part.id == part_id).first()
    if not part:
        raise HTTPException(status_code=404, detail="Part not found")

    current_images = part.images or []
    if len(current_images) >= MAX_IMAGES:
        raise HTTPException(status_code=400, detail=f"Max {MAX_IMAGES} images per part")

    if file.content_type not in ALLOWED_TYPES:
        raise HTTPException(status_code=400, detail="Only JPEG, PNG and WebP are allowed")

    content = await file.read()
    if len(content) > MAX_FILE_SIZE_MB * 1024 * 1024:
        raise HTTPException(status_code=400, detail=f"File too large (max {MAX_FILE_SIZE_MB} MB)")

    os.makedirs(UPLOAD_DIR, exist_ok=True)
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else "jpg"
    filename = f"{part_id}_{uuid.uuid4().hex[:8]}.{ext}"
    filepath = os.path.join(UPLOAD_DIR, filename)

    # Lagre og resize med Pillow (maks 1200px bredde, behold ratio)
    import io
    img = PilImage.open(io.BytesIO(content))
    img = img.convert("RGB")
    if img.width > 1200:
        ratio = 1200 / img.width
        img = img.resize((1200, int(img.height * ratio)), PilImage.LANCZOS)
    img.save(filepath, optimize=True, quality=85)

    part.images = current_images + [f"/uploads/{filename}"]
    db.commit()
    db.refresh(part)
    return part


# ─── Slett ett bilde ──────────────────────────────────────────────────
@router.delete("/{part_id}/images/{image_index}", response_model=PartOut)
def delete_image(
    part_id: int,
    image_index: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    part = db.query(Part).filter(Part.id == part_id).first()
    if not part:
        raise HTTPException(status_code=404, detail="Part not found")

    images = part.images or []
    if image_index < 0 or image_index >= len(images):
        raise HTTPException(status_code=404, detail="Image not found")

    img_to_delete = images[image_index]
    full_path = os.path.join(UPLOAD_DIR, os.path.basename(img_to_delete))
    if os.path.exists(full_path):
        os.remove(full_path)

    part.images = [img for i, img in enumerate(images) if i != image_index]
    db.commit()
    db.refresh(part)
    return part


# ─── Bulk import from Excel / CSV ─────────────────────────────────────
# Accepts flexible column names (English or Norwegian). Required columns:
# part number, name, category, price. Everything else is optional.
COLUMN_ALIASES = {
    "part_number": ["part_number", "part number", "part no", "part no.", "partno", "varenummer", "varenr", "vare nr", "sku", "artikkelnr"],
    "name": ["name", "part name", "navn", "varenavn", "produktnavn"],
    "category": ["category", "kategori"],
    "unit_price": ["unit_price", "price", "price (nok)", "pris", "unit price", "salgspris"],
    "description": ["description", "beskrivelse", "desc"],
    "stock_quantity": ["stock_quantity", "stock", "stock quantity", "antall", "lager", "quantity", "antall pa lager", "antall på lager"],
    "low_stock_threshold": ["low_stock_threshold", "threshold", "low stock threshold", "lavt lager", "min", "minimum"],
    "location": ["location", "storage location", "lokasjon", "hylle", "plassering", "hylleplass"],
    "ordered_quantity": ["ordered_quantity", "ordered", "ordered quantity", "bestilt", "i bestilling"],
}


def _opt_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return None if s == "" or s.lower() == "nan" else s


def _opt_int(v, default):
    try:
        if v is None or str(v).strip().lower() in ("", "nan"):
            return default
        return int(float(v))
    except (ValueError, TypeError):
        return default


@router.post("/import")
async def import_parts(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Only admins can import parts")

    import io
    import pandas as pd

    content = await file.read()
    fname = (file.filename or "").lower()
    try:
        if fname.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(content))
        else:
            df = pd.read_excel(io.BytesIO(content))
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not read the file: {exc}")

    # Map the user's column headers onto our field names
    rename = {}
    for col in df.columns:
        key = str(col).strip().lower()
        for field, aliases in COLUMN_ALIASES.items():
            if key in aliases:
                rename[col] = field
                break
    df = df.rename(columns=rename)

    # Only the part number is required. Name, category and price can be filled
    # in later, so a row with just a part number (and maybe a stock count) is
    # enough to get it into the system.
    if "part_number" not in df.columns:
        raise HTTPException(
            status_code=400,
            detail="The file needs a column for the part number (e.g. 'part_number' or 'varenummer').",
        )

    existing = {r[0] for r in db.query(Part.part_number).all()}
    created, skipped, errors, seen = 0, 0, [], set()

    for i, row in df.iterrows():
        rownum = int(i) + 2  # +1 for header, +1 for 1-based
        try:
            pn = _opt_str(row.get("part_number"))
            if not pn:
                errors.append({"row": rownum, "reason": "missing part number"}); continue
            if pn in existing or pn in seen:
                skipped += 1; continue
            # Fill in sensible placeholders for anything left blank – the shop
            # can complete name, category and price later by editing the part.
            name = _opt_str(row.get("name")) or pn        # fall back to the part number
            category = _opt_str(row.get("category")) or "Uncategorized"
            try:
                price = float(row.get("unit_price"))
                if price != price:                        # NaN (blank Excel cell) → unknown
                    price = 0.0
            except (ValueError, TypeError):
                price = 0.0                                # unknown price for now

            db.add(Part(
                part_number=pn, name=name, category=category, unit_price=price,
                description=_opt_str(row.get("description")),
                location=_opt_str(row.get("location")),
                stock_quantity=_opt_int(row.get("stock_quantity"), 0),
                ordered_quantity=_opt_int(row.get("ordered_quantity"), 0),
                low_stock_threshold=_opt_int(row.get("low_stock_threshold"), 5),
                compatible_cars=[], images=[],
            ))
            seen.add(pn)
            created += 1
        except Exception as exc:
            errors.append({"row": rownum, "reason": str(exc)})

    db.commit()
    return {
        "created": created,
        "skipped_duplicates": skipped,
        "error_count": len(errors),
        "errors": errors[:50],
    }
